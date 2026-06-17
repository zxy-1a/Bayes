from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import io
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

# 淇 Windows 缂栫爜闂
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors='replace')


SEP_RE = re.compile(r"[\u3001\uff0c,;\uff1b\u3002\r\n]+")
SPACE_RE = re.compile(r"\s+")
COMPARE_RE = re.compile(r"\s*/\s*")
TRIM_RE = re.compile(r"^[\u6709\u60a3]|\u7b49$|\u7b49\u75c7$|\u75c7$|\u75c5$")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u2022", "").replace("\r", "").strip()
    return text


def norm_key(value: object) -> str:
    text = clean_text(value)
    text = SPACE_RE.sub("", text)
    text = text.strip("\u3001\uff0c,\uff1b;\u3002.:\uff1a")
    return text.lower()


def term_variants(term: str) -> set[str]:
    term = norm_key(term)
    if not term:
        return set()

    variants = {term}
    trimmed = TRIM_RE.sub("", term)
    if trimmed and trimmed != term:
        variants.add(trimmed)

    for prefix in ("\u7537\u6027", "\u5973\u6027", "\u5404\u7c7b"):
        if term.startswith(prefix) and len(term) > len(prefix) + 1:
            variants.add(term[len(prefix):])

    replacements = {
        "\u6025\u6162\u6027": ["\u6025\u6027", "\u6162\u6027", ""],
        "\u504f\u9ad8": ["\u9ad8"],
        "\u4e0d\u8db3": ["\u865a"],
    }
    for old, news in replacements.items():
        if old in term:
            for new in news:
                variant = term.replace(old, new)
                if len(variant) >= 2:
                    variants.add(variant)

    return {v for v in variants if len(v) >= 2}


def split_terms(value: object) -> list[str]:
    text = clean_text(value)
    # Excel line breaks sometimes split one word, e.g. "\u5c0f\n\u513f\u591a\u52a8".
    text = text.replace("\n", "")
    terms = []
    for part in SEP_RE.split(text):
        part = norm_key(part)
        if part:
            terms.append(part)
    return terms


def first_non_empty(values: Iterable[object]) -> str:
    for value in values:
        text = norm_key(value)
        if text:
            return text
    return ""


def tea_list(value: object) -> list[str]:
    text = clean_text(value)
    return [norm_key(x) for x in re.split(r"[\s+銆侊紝,;锛沑n]+", text) if norm_key(x)]


def append_alias(
    rows: list[dict[str, object]],
    alias: object,
    canonical: object,
    source: str,
    confidence: float,
) -> None:
    alias_text = norm_key(alias)
    canonical_text = norm_key(canonical)
    if not alias_text or not canonical_text or alias_text.isdigit():
        return
    if len(alias_text) > 60 or len(canonical_text) > 60:
        return

    for alias_variant in term_variants(alias_text):
        for canonical_variant in term_variants(canonical_text):
            rows.append(
                {
                    "alias": alias_variant,
                    "canonical": canonical_variant,
                    "source": source,
                    "confidence": confidence,
                }
            )


def extract_definition_aliases(definition: str) -> list[str]:
    text = clean_text(definition)
    aliases: list[str] = []

    if "\uff1a" in text:
        head, tail = text.split("\uff1a", 1)
        aliases.extend(split_terms(head))
        first_sentence = re.split(r"[\u3002.]", tail, maxsplit=1)[0]
        aliases.extend(split_terms(first_sentence))

    for marker in ("\u53c8\u79f0", "\u53c8\u540d", "\u4ea6\u79f0"):
        if marker in text:
            tail = text.split(marker, 1)[1]
            tail = re.split(r"[\u3002.]", tail, maxsplit=1)[0]
            aliases.extend(split_terms(tail))

    return [x for x in aliases if 2 <= len(x) <= 20]


def choose_xlsx_files(base_dir: Path) -> tuple[Path, Path]:
    search_dirs = [base_dir, base_dir / "inputs"]
    files = sorted(
        [
            p
            for search_dir in search_dirs
            if search_dir.exists()
            for p in search_dir.glob("*.xlsx")
            if not p.name.startswith("~$")
        ],
        key=lambda p: p.stat().st_size,
    )
    if len(files) < 2:
        raise FileNotFoundError("Need at least two .xlsx files in the working directory.")
    # The tea matching table is tiny; the terminology workbook is much larger.
    return files[0], files[-1]


def load_tea_rows(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    tea_rows: list[dict[str, object]] = []
    symptom_index: list[dict[str, object]] = []
    current_combo_rules = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tea_name = norm_key(row[0])
        if not tea_name:
            continue

        course = norm_key(row[1])
        raw_symptom = clean_text(row[2])
        combo_rules = clean_text(row[3]) or current_combo_rules
        if clean_text(row[3]):
            current_combo_rules = clean_text(row[3])

        constitution = norm_key(row[4])
        constitution_teas = tea_list(row[5])
        organ = norm_key(row[6])
        organ_teas = tea_list(row[7])
        symptom_terms = split_terms(raw_symptom)

        tea_rows.append(
            {
                "tea_name": tea_name,
                "course": course,
                "raw_symptom": raw_symptom,
                "symptom_terms": symptom_terms,
                "combo_rules": combo_rules,
                "constitution": constitution,
                "constitution_teas": constitution_teas,
                "organ": organ,
                "organ_teas": organ_teas,
                "source_file": path.name,
                "source_row": row_idx,
            }
        )

        for symptom in symptom_terms:
            symptom_index.append(
                {
                    "symptom": symptom,
                    "tea_name": tea_name,
                    "course": course,
                    "source_row": row_idx,
                    "raw_symptom": raw_symptom,
                }
            )

    return tea_rows, symptom_index


def load_alias_rows(term_path: Path, tea_terms: Iterable[str]) -> list[dict[str, object]]:
    wb = load_workbook(term_path, data_only=True)
    alias_rows: list[dict[str, object]] = []

    for ws in wb.worksheets:
        if ws.title == "Sheet1":
            for row in ws.iter_rows(values_only=True):
                canonical = first_non_empty([row[3] if len(row) > 3 else "", row[1] if len(row) > 1 else ""])
                if not canonical:
                    continue
                append_alias(alias_rows, canonical, canonical, "term_sheet1", 1.0)
                if len(row) > 5 and row[5]:
                    for alias in extract_definition_aliases(str(row[5])):
                        append_alias(alias_rows, alias, canonical, "term_sheet1_definition", 0.75)

        elif ws.title == "Sheet2":
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Header: original, unified, merged, split, compared_standard, standard, alias1.
                canonical = first_non_empty(
                    [
                        row[5] if len(row) > 5 else "",
                        row[4] if len(row) > 4 else "",
                        row[1] if len(row) > 1 else "",
                        row[3] if len(row) > 3 else "",
                        row[0] if len(row) > 0 else "",
                    ]
                )
                if "/" in canonical:
                    canonical = COMPARE_RE.split(canonical)[-1]

                for idx in (0, 1, 3, 4, 5, 6):
                    if idx < len(row) and row[idx]:
                        for alias in split_terms(row[idx]):
                            append_alias(alias_rows, alias, canonical, "term_sheet2", 1.0)

        elif ws.title in {"Sheet3", "Sheet4"}:
            start_row = 2 if ws.title == "Sheet3" else 1
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                for value in row:
                    for term in split_terms(value):
                        append_alias(alias_rows, term, term, f"term_{ws.title.lower()}_self", 0.6)

    # Product matching terms are business-critical even when the clinical term file lacks them.
    for term in tea_terms:
        append_alias(alias_rows, term, term, "tea_matching_term", 0.95)

    dedup: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in alias_rows:
        key = (str(row["alias"]), str(row["canonical"]), str(row["source"]))
        if key not in dedup or float(row["confidence"]) > float(dedup[key]["confidence"]):
            dedup[key] = row

    return sorted(dedup.values(), key=lambda r: (str(r["alias"]), str(r["canonical"]), str(r["source"])))


def split_combo_rules(raw_rules: str) -> list[str]:
    text = clean_text(raw_rules)
    if not text:
        return []
    text = text.replace("\u2022", "\n")
    text = re.sub(r"\s*(\d+[、.．])", r"\n\1", text)
    parts = re.split(r"(?:^|\n)\d+[、.．]", text)
    return [part.strip("：:；;、，,") for part in parts if part.strip("：:；;、，,")]


def extract_teas(value: str, known_teas: list[str]) -> list[str]:
    text = norm_key(value)
    hits: list[tuple[int, str]] = []
    for tea in known_teas:
        pos = text.find(tea)
        if pos >= 0:
            hits.append((pos, tea))

    if hits:
        seen: set[str] = set()
        ordered = []
        for _, tea in sorted(hits, key=lambda item: item[0]):
            if tea not in seen:
                seen.add(tea)
                ordered.append(tea)
        return ordered

    teas = []
    for token in re.split(r"[+\s、，,;；]+", text):
        if token.endswith("茶") and 2 <= len(token) <= 5:
            teas.append(token)
    return teas


def first_tea_pos(value: str, known_teas: list[str]) -> int:
    text = norm_key(value)
    positions = [text.find(tea) for tea in known_teas if text.find(tea) >= 0]
    return min(positions) if positions else -1


def parse_combo_rule(rule_text: str, rule_id: int, known_teas: list[str]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not rule_text or ("：" not in rule_text and ":" not in rule_text):
        return rows

    condition_part, recommendation_part = re.split(r"[：:]", rule_text, maxsplit=1)
    primary_conditions = split_terms(condition_part)
    fragments = re.split(r"[；;]", recommendation_part)

    for fragment_idx, fragment in enumerate(fragments, start=1):
        fragment = norm_key(fragment)
        if not fragment or "+" not in fragment:
            continue

        if "：" in fragment or ":" in fragment:
            trigger, teas_text = re.split(r"[：:]", fragment, maxsplit=1)
        else:
            pos = first_tea_pos(fragment, known_teas)
            trigger = fragment[:pos] if pos > 0 else ""
            teas_text = fragment[pos:] if pos > 0 else fragment

        teas = extract_teas(teas_text, known_teas)
        if not teas:
            continue

        add_on_teas: list[str] = []
        base_teas = teas
        if trigger.startswith("如有") or "如有" in trigger:
            add_on_teas = teas[-1:]
            base_teas = teas[:-1] or teas

        rows.append(
            {
                "rule_id": f"step2::{rule_id}.{fragment_idx}",
                "match_stage": "step2_symptom_combination",
                "primary_conditions": "、".join(primary_conditions),
                "trigger_condition": trigger,
                "recommended_teas": "、".join(teas),
                "base_teas": "、".join(base_teas),
                "add_on_teas": "、".join(add_on_teas),
                "priority": 2,
                "rule_type": "add_on" if add_on_teas else "combination",
                "raw_rule": rule_text,
            }
        )

    return rows


def build_combo_rule_rows(tea_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    combo_rules = first_non_empty(row["combo_rules"] for row in tea_rows)
    known_teas = sorted({str(row["tea_name"]) for row in tea_rows}, key=len, reverse=True)
    rows: list[dict[str, object]] = []
    for idx, rule in enumerate(split_combo_rules(combo_rules), start=1):
        rows.extend(parse_combo_rule(rule, idx, known_teas))
    return rows


def build_constitution_fallback_rows(tea_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for row in tea_rows:
        constitution = str(row["constitution"])
        if not constitution:
            continue
        for tea in row["constitution_teas"]:
            key = (constitution, tea)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "fallback_type": "constitution",
                    "fallback_condition": constitution,
                    "tea_name": tea,
                    "priority": 3,
                    "weight": 0.6,
                    "applies_when": "no_step1_or_step2_match",
                    "source_row": row["source_row"],
                }
            )
    return rows


def build_organ_fallback_rows(tea_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for row in tea_rows:
        organ = str(row["organ"])
        if not organ:
            continue
        for tea in row["organ_teas"]:
            key = (organ, tea)
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "fallback_type": "organ",
                    "fallback_condition": organ,
                    "tea_name": tea,
                    "priority": 4,
                    "weight": 0.6,
                    "applies_when": "no_step1_or_step2_match",
                    "source_row": row["source_row"],
                }
            )
    return rows


def build_decision_policy() -> list[dict[str, object]]:
    return [
        {
            "stage": "step1_symptom_initial_screen",
            "priority": 1,
            "action": "Use tea_symptom_index and exact/alias matches first. These are primary recommendations.",
            "weight": 1.0,
        },
        {
            "stage": "step2_symptom_combination",
            "priority": 2,
            "action": "If a combination rule is matched, recommend the rule teas. Add-on teas are included only when their trigger condition is present.",
            "weight": 1.2,
        },
        {
            "stage": "step3_constitution_fallback",
            "priority": 3,
            "action": "Use constitution fallback only when step1 and step2 do not match any user complaint.",
            "weight": 0.6,
        },
        {
            "stage": "step4_organ_fallback",
            "priority": 4,
            "action": "Use organ fallback together with constitution fallback only when step1 and step2 do not match. If the same tea appears in both, increase its fallback rank.",
            "weight": 0.6,
        },
    ]


def build_rag_docs(
    tea_rows: list[dict[str, object]],
    combo_rule_rows: list[dict[str, object]],
    constitution_fallback_rows: list[dict[str, object]],
    organ_fallback_rows: list[dict[str, object]],
    decision_policy_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    docs: list[dict[str, object]] = []

    for row in tea_rows:
        tea_name = str(row["tea_name"])
        symptom_terms = list(row["symptom_terms"])
        content_parts = [
            f"tea_name: {tea_name}",
            "match_stage: step1_symptom_initial_screen",
            "priority: 1",
            f"recommended_course: {row['course']}",
            "matched_symptoms: " + "\u3001".join(symptom_terms),
            f"raw_matching_logic: {row['raw_symptom']}",
        ]
        if row["constitution"]:
            content_parts.append(f"constitution_filter: {row['constitution']}")
        if row["organ"]:
            content_parts.append(f"organ_filter: {row['organ']}")

        docs.append(
            {
                "id": f"tea::{tea_name}",
                "content": "\n".join(content_parts),
                "metadata": {
                    "doc_type": "tea_product",
                    "match_stage": "step1_symptom_initial_screen",
                    "priority": 1,
                    "tea_name": tea_name,
                    "course": row["course"],
                    "symptom_terms": symptom_terms,
                    "constitution": row["constitution"],
                    "organ": row["organ"],
                    "source_file": row["source_file"],
                    "source_row": row["source_row"],
                },
            }
        )

        for symptom in symptom_terms:
            docs.append(
                {
                    "id": f"symptom::{symptom}::{tea_name}",
                    "content": (
                        f"customer_symptom_or_disease: {symptom}\n"
                        "match_stage: step1_symptom_initial_screen\n"
                        "priority: 1\n"
                        f"recommend_tea: {tea_name}\n"
                        f"recommended_course: {row['course']}\n"
                        f"evidence: {row['raw_symptom']}"
                    ),
                    "metadata": {
                        "doc_type": "symptom_to_tea",
                        "match_stage": "step1_symptom_initial_screen",
                        "priority": 1,
                        "symptom": symptom,
                        "tea_name": tea_name,
                        "course": row["course"],
                        "source_file": row["source_file"],
                        "source_row": row["source_row"],
                    },
                }
            )

    combo_rules = first_non_empty(row["combo_rules"] for row in tea_rows)
    if combo_rules:
        docs.append(
            {
                "id": "rule::combination",
                "content": (
                    "match_stage: step2_symptom_combination\n"
                    "priority: 2\n"
                    f"combination_recommendation_rules:\n{combo_rules}"
                ),
                "metadata": {"doc_type": "combination_rule", "match_stage": "step2_symptom_combination", "priority": 2},
            }
        )

    for row in combo_rule_rows:
        rule_id = str(row["rule_id"])
        content_parts = [
            f"rule_id: {rule_id}",
            "match_stage: step2_symptom_combination",
            f"priority: {row['priority']}",
            f"primary_conditions: {row['primary_conditions']}",
            f"recommended_teas: {row['recommended_teas']}",
            f"base_teas: {row['base_teas']}",
            f"rule_type: {row['rule_type']}",
            "decision: Step2 is a primary matching stage. Apply this rule when the user complaint matches the primary condition.",
        ]
        if row["trigger_condition"]:
            content_parts.append(f"trigger_condition: {row['trigger_condition']}")
        if row["add_on_teas"]:
            content_parts.append(f"add_on_teas: {row['add_on_teas']}")
            content_parts.append("add_on_policy: Add these teas only when the trigger condition is present.")

        docs.append(
            {
                "id": f"rule::{rule_id}",
                "content": "\n".join(content_parts),
                "metadata": {
                    "doc_type": "structured_combo_rule",
                    "match_stage": "step2_symptom_combination",
                    "priority": int(row["priority"]),
                    "rule_id": rule_id,
                    "primary_conditions": str(row["primary_conditions"]).split("、"),
                    "trigger_condition": row["trigger_condition"],
                    "recommended_teas": str(row["recommended_teas"]).split("、"),
                    "base_teas": str(row["base_teas"]).split("、"),
                    "add_on_teas": str(row["add_on_teas"]).split("、") if row["add_on_teas"] else [],
                    "rule_type": row["rule_type"],
                },
            }
        )

    for row in constitution_fallback_rows:
        docs.append(
            {
                "id": f"fallback::constitution::{row['fallback_condition']}::{row['tea_name']}",
                "content": (
                    "match_stage: step3_constitution_fallback\n"
                    f"fallback_condition: {row['fallback_condition']}\n"
                    f"recommend_tea: {row['tea_name']}\n"
                    f"priority: {row['priority']}\n"
                    f"weight: {row['weight']}\n"
                    "applies_when: only when Step1 and Step2 have no related match"
                ),
                "metadata": {
                    "doc_type": "fallback_rule",
                    "fallback_type": "constitution",
                    "match_stage": "step3_constitution_fallback",
                    "priority": int(row["priority"]),
                    "fallback_condition": row["fallback_condition"],
                    "tea_name": row["tea_name"],
                    "weight": row["weight"],
                    "applies_when": row["applies_when"],
                    "source_row": row["source_row"],
                },
            }
        )

    for row in organ_fallback_rows:
        docs.append(
            {
                "id": f"fallback::organ::{row['fallback_condition']}::{row['tea_name']}",
                "content": (
                    "match_stage: step4_organ_fallback\n"
                    f"fallback_condition: {row['fallback_condition']}\n"
                    f"recommend_tea: {row['tea_name']}\n"
                    f"priority: {row['priority']}\n"
                    f"weight: {row['weight']}\n"
                    "applies_when: only when Step1 and Step2 have no related match. If the same tea also appears in Step3 fallback, raise its fallback rank."
                ),
                "metadata": {
                    "doc_type": "fallback_rule",
                    "fallback_type": "organ",
                    "match_stage": "step4_organ_fallback",
                    "priority": int(row["priority"]),
                    "fallback_condition": row["fallback_condition"],
                    "tea_name": row["tea_name"],
                    "weight": row["weight"],
                    "applies_when": row["applies_when"],
                    "source_row": row["source_row"],
                },
            }
        )

    for row in decision_policy_rows:
        docs.append(
            {
                "id": f"policy::{row['stage']}",
                "content": (
                    f"match_stage: {row['stage']}\n"
                    f"priority: {row['priority']}\n"
                    f"weight: {row['weight']}\n"
                    f"policy_action: {row['action']}"
                ),
                "metadata": {
                    "doc_type": "matching_policy_stage",
                    "match_stage": row["stage"],
                    "priority": int(row["priority"]),
                    "weight": row["weight"],
                },
            }
        )

    docs.append(
        {
            "id": "policy::matching_priority",
            "content": (
                "matching_priority_policy:\n"
                "1. Step1 symptom initial screen and Step2 symptom combination are primary.\n"
                "2. If Step1/Step2 match user complaint, do not use Step3/Step4 to override them.\n"
                "3. Only when Step1/Step2 find no related complaint, use Step3 constitution and Step4 organ fallback.\n"
                "4. In fallback mode, teas appearing in both constitution and organ candidates receive higher rank.\n"
                "5. Add-on rules such as respiratory disease with qi deficiency add Yuanqi Tea only when the add-on trigger is present."
            ),
            "metadata": {"doc_type": "matching_policy", "priority": 0},
        }
    )

    return docs


def find_unmatched(
    symptom_index: list[dict[str, object]],
    alias_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    aliases = {str(row["alias"]) for row in alias_rows if str(row["source"]) != "tea_matching_term"}
    unmatched = []
    for row in symptom_index:
        symptom = str(row["symptom"])
        variants = term_variants(symptom)
        if not (variants & aliases):
            unmatched.append(row)
    return unmatched


WESTERN_DISEASE_TERMS = {
    "高血压",
    "高血脂",
    "高血糖",
    "高尿酸",
    "流感",
    "咽炎",
    "支气管炎",
    "慢性支气管炎",
    "肺气肿",
    "肺结节",
    "眼底血管病",
    "糖尿病",
    "痛风",
    "肾结石",
    "肾炎",
    "肾功能不全",
    "急慢性胃炎",
    "十二指肠球部溃疡",
    "肠化生",
    "反流性食道炎",
    "胃酸过多",
    "急慢性肠炎",
    "记忆力减退",
    "注意力不集中",
    "自闭症",
    "老年痴呆",
    "偏执型人格障碍",
    "脂肪肝",
    "肝炎",
    "胆结石",
    "心肌炎",
    "心肌炎后遗症",
    "冠心病",
    "各类结节病",
    "癌前病变",
    "子宫肌瘤",
    "卵巢囊肿",
    "乳腺结节",
    "甲状腺结节",
    "体重超标",
}


BUSINESS_ALIAS_TERMS = {
    "改善心肺功能弱",
    "易疲劳",
    "有湿气",
    "手脚不温",
    "护肝明目",
    "缓解眼疲劳",
    "抗病毒",
    "抗脂肪化",
    "抗纤维化",
    "护肝利胆",
    "腰酸腿软",
    "女性宫寒",
    "月经不调",
    "痛经",
    "面容晦暗",
    "心脏亚健康",
}


FUNCTION_COPY_TERMS = {
    "护肝明目",
    "缓解眼疲劳",
    "抗病毒",
    "抗脂肪化",
    "抗纤维化",
    "护肝利胆",
}


SYMPTOM_STATE_TERMS = {
    "改善心肺功能弱",
    "易疲劳",
    "有湿气",
    "手脚不温",
    "腰酸腿软",
    "面容晦暗",
    "心脏亚健康",
}


WESTERN_KEYWORDS = (
    "高血压", "高血脂", "高血糖", "高尿酸", "流感", "咽炎", "支气管炎", "肺", "胃", "肠", "肝", "肾",
    "结节", "糖尿病", "心肌", "冠心", "炎", "癌", "肌瘤", "囊肿", "痴呆", "抑郁", "肥胖",
)

BUSINESS_KEYWORDS = (
    "护肝", "缓解", "抗", "改善", "疲劳", "湿气", "不温", "腰酸", "宫寒", "不调", "痛经", "晦暗", "亚健康", "体重超标",
)


def classify_unmatched_term(term: str) -> tuple[str, str, str]:
    t = norm_key(term)
    if t in WESTERN_DISEASE_TERMS:
        return "western_disease", "disease_or_finding", "direct medical label"
    if t in BUSINESS_ALIAS_TERMS:
        if t in FUNCTION_COPY_TERMS:
            return "business_alias", "function_copy", "marketing/functional wording"
        if t in SYMPTOM_STATE_TERMS:
            return "business_alias", "symptom_state", "oral symptom/state wording"
        return "business_alias", "tcm_state", "TCM or condition wording"

    if any(keyword in t for keyword in WESTERN_KEYWORDS):
        return "western_disease", "disease_or_finding", "keyword hit"

    if any(keyword in t for keyword in BUSINESS_KEYWORDS):
        return "business_alias", "symptom_state", "keyword hit"

    return "business_alias", "needs_review", "manual check"

def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-dir", type=Path, default=Path.cwd())
    parser.add_argument("--tea-xlsx", type=Path)
    parser.add_argument("--term-xlsx", type=Path)
    parser.add_argument("--out-dir", type=Path, default=Path("rag_output"))
    args = parser.parse_args()

    tea_path, term_path = (
        (args.tea_xlsx, args.term_xlsx)
        if args.tea_xlsx and args.term_xlsx
        else choose_xlsx_files(args.base_dir)
    )
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    tea_rows, symptom_index = load_tea_rows(tea_path)
    alias_rows = load_alias_rows(term_path, [row["symptom"] for row in symptom_index])
    combo_rule_rows = build_combo_rule_rows(tea_rows)
    constitution_fallback_rows = build_constitution_fallback_rows(tea_rows)
    organ_fallback_rows = build_organ_fallback_rows(tea_rows)
    decision_policy_rows = build_decision_policy()
    rag_docs = build_rag_docs(
        tea_rows,
        combo_rule_rows,
        constitution_fallback_rows,
        organ_fallback_rows,
        decision_policy_rows,
    )
    unmatched = find_unmatched(symptom_index, alias_rows)
    classified_unmatched = []
    for row in unmatched:
        category, subtype, note = classify_unmatched_term(str(row["symptom"]))
        classified_unmatched.append(
            {
                **row,
                "category": category,
                "subtype": subtype,
                "note": note,
            }
        )

    write_jsonl(out_dir / "tea_rag_documents.jsonl", rag_docs)
    write_csv(out_dir / "tea_symptom_index.csv", symptom_index)
    write_csv(out_dir / "step2_combo_rules.csv", combo_rule_rows)
    write_csv(out_dir / "step3_constitution_fallback.csv", constitution_fallback_rows)
    write_csv(out_dir / "step4_organ_fallback.csv", organ_fallback_rows)
    write_csv(out_dir / "matching_decision_policy.csv", decision_policy_rows)
    write_csv(out_dir / "symptom_alias_dictionary.csv", alias_rows)
    write_csv(out_dir / "unmatched_tea_terms.csv", unmatched)
    write_csv(out_dir / "unmatched_tea_terms_classified.csv", classified_unmatched)

    western_alias = [row for row in classified_unmatched if row["category"] == "western_disease"]
    business_alias = [row for row in classified_unmatched if row["category"] == "business_alias"]
    write_csv(out_dir / "western_disease_alias.csv", western_alias)
    write_csv(out_dir / "business_alias.csv", business_alias)

    source_summary = [
        {"item": "tea_file", "value": str(tea_path)},
        {"item": "term_file", "value": str(term_path)},
        {"item": "tea_count", "value": len(tea_rows)},
        {"item": "tea_symptom_link_count", "value": len(symptom_index)},
        {"item": "step2_combo_rule_count", "value": len(combo_rule_rows)},
        {"item": "step3_constitution_fallback_count", "value": len(constitution_fallback_rows)},
        {"item": "step4_organ_fallback_count", "value": len(organ_fallback_rows)},
        {"item": "alias_count", "value": len(alias_rows)},
        {"item": "rag_doc_count", "value": len(rag_docs)},
        {"item": "unmatched_tea_term_count", "value": len(unmatched)},
        {"item": "western_disease_alias_count", "value": len(western_alias)},
        {"item": "business_alias_count", "value": len(business_alias)},
    ]
    write_csv(out_dir / "summary.csv", source_summary)

    print(json.dumps(source_summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


